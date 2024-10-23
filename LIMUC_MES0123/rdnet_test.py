import torch
import torch.nn as nn
from torch.utils.data import DataLoader
from torchvision import datasets, transforms
import timm
from util.logger_util import write_test_data_log
import json
from  util.logger_util import setup_logger
import os
from openpyxl import Workbook,load_workbook


def main():
    log_file = os.path.join(os.getcwd(), "test_log.txt")
    logger = setup_logger(log_file)
    logger.info("-----------------------------")
    # 设置设备

    k_fold = 4
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    model_name = 'rdnet_base.nv_in1k'
    model_file = f'./output/best_model_{k_fold}.pth'

    excel_file = 'test_log.xlsx'
    sheet_name = f'fold {k_fold}'
    excel_file_exists = os.path.exists(excel_file)

    # 创建或加载工作簿
    if excel_file_exists:
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    ws.append(["image", "True label", "Predicted label", "Probability"])

    # 数据预处理
    transform_test = transforms.Compose([
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ])

    # 加载测试数据集
    test_dataset = datasets.ImageFolder(root=r'../dataset/LIMUCMerged/test', transform=transform_test)
    test_loader = DataLoader(test_dataset, batch_size=32, shuffle=False, num_workers=0)

    # 加载模型
    model = timm.create_model(model_name, pretrained=False, num_classes=len(test_dataset.classes))
    model.load_state_dict(torch.load(model_file))
    model.to(device)
    model.eval()  # 设置模型为评估模式


    class_names_dict = {str(i): name for i, name in enumerate(test_dataset.classes)}
    classes_json = json.dumps(class_names_dict, indent=4)
 #   print(f'Classes in JSON format: \n{classes_json}')

    results=[]
    # 测试模型
    total_correct = 0
    with torch.no_grad():
        for inputs, labels in test_loader:
            inputs, labels = inputs.to(device), labels.to(device)
            outputs = model(inputs)
            probabilities = torch.softmax(outputs, dim=1)  # 计算预测概率
            _, predicted = torch.max(outputs.data, 1)

            # 将真实标签、预测标签和预测概率添加到结果列表中
            for i in range(len(labels)):
                true_label = labels[i].item()
                pred_label = predicted[i].item()
                pred_prob = probabilities[i][pred_label].item()
               # results.append((true_label, pred_label, pred_prob))
                results.append((true_label, pred_label, pred_prob, probabilities[i][0].item(),probabilities[i][1].item(),probabilities[i][2].item(),probabilities[i][3].item()))

            total_correct += (predicted == labels).sum().item()

    test_accuracy = total_correct / len(test_dataset)
    print(f'Test Accuracy: {test_accuracy:.4f}')

    msg = f'{model_name}, {model_file},Test Accuracy: {test_accuracy:.4f}'
    logger.info(msg)
    # 打印或保存结果
    for idx, (true_label, pred_label, pred_prob, pred_prob0, pred_prob1, pred_prob2, pred_prob3) in enumerate(results):

        logger.info(f'Image {idx}: True label: {true_label}, Predicted label: {pred_label}, Probability: {pred_prob:.4f}')
        print(f'Image {idx}: True label: {true_label}, Predicted label: {pred_label}, Probability: {pred_prob:.4f}')
        #ws.append([idx, true_label, pred_label,pred_prob])
        ws.append([idx, true_label, pred_label,pred_prob,pred_prob0,pred_prob1,pred_prob2,pred_prob3])

    write_test_data_log(model_name, model_file, test_accuracy)
    print(f'Test Accuracy: {test_accuracy:.4f}')
    wb.save(excel_file)


if __name__ == '__main__':
    main()